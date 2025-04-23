import pandas as pd
import random
from datetime import datetime, time, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from collections import defaultdict

# Constants
DAYS = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
START_TIME = time(9, 0)
END_TIME = time(18, 30)
LECTURE_DURATION = 3  # 1.5 hours = 3 slots (30 mins each)
LAB_DURATION = 4      # 2 hours = 4 slots (30 mins each)
TUTORIAL_DURATION = 2  # 30 mins = 1 slot

# Color palette for different courses (pastel colors)
COLOR_PALETTE = [
    "FFD6E0", "FFEFCF", "D6FFCF", "CFFEFF", "D6CFFF", "FFCFF4", 
    "E8D0A9", "B7E1CD", "C9DAF8", "FFD6CC", "D9D2E9", "EAD1DC",
    "A4C2F4", "D5A6BD", "B6D7A8", "FFE599", "A2C4C9", "D5D5D5"
]

# Load room data
try:
    rooms_df = pd.read_csv('Rooms.csv')
    # Create separate lists for lecture rooms and lab rooms
    lecture_rooms = rooms_df[rooms_df['type'] == 'LECTURE_ROOM']['roomNumber'].tolist()
    computer_lab_rooms = rooms_df[rooms_df['type'] == 'COMPUTER_LAB']['roomNumber'].tolist()
    large_rooms = rooms_df[rooms_df['type'] == 'SEATER_120']['roomNumber'].tolist()
    
    # Add error handling for room types
    if not lecture_rooms:
        print("Warning: No LECTURE_ROOM type rooms found in Rooms.csv")
    if not computer_lab_rooms:
        print("Warning: No COMPUTER_LAB type rooms found in Rooms.csv")
    if not large_rooms:
        print("Warning: No SEATER_120 type rooms found in Rooms.csv")
except FileNotFoundError:
    print("Error: File 'Rooms.csv' not found in the current directory")
    lecture_rooms = []
    computer_lab_rooms = []
    large_rooms = []
except Exception as e:
    print(f"Error loading Rooms.csv: {e}")
    lecture_rooms = []
    computer_lab_rooms = []
    large_rooms = []

def generate_course_color():
    """Generate unique colors for courses from the palette or random if needed"""
    for color in COLOR_PALETTE:
        yield color
    
    # If we run out of predefined colors, generate random ones
    while True:
        r = format(random.randint(180, 255), '02x')
        g = format(random.randint(180, 255), '02x')
        b = format(random.randint(180, 255), '02x')
        yield f"{r}{g}{b}"

def generate_time_slots():
    slots = []
    current_time = datetime.combine(datetime.today(), START_TIME)
    end_time = datetime.combine(datetime.today(), END_TIME)
    
    while current_time < end_time:
        current = current_time.time()
        next_time = current_time + timedelta(minutes=30)
        
        # Keep all time slots but we'll mark break times later
        slots.append((current, next_time.time()))
        current_time = next_time
    
    return slots

# Load data from Excel
try:
    df = pd.read_excel('combined.xlsx', sheet_name='Sheet1')
except FileNotFoundError:
    print("Error: File 'combined.xlsx' not found in the current directory")
    exit()

def is_break_time(slot):
    """Check if a time slot falls within break times"""
    start, end = slot
    # Morning break: 10:30-11:00 (15 minutes)
    morning_break = (time(10, 30) <= start < time(10, 45))
    # Lunch break: 13:00-14:00 (45 minutes)
    lunch_break = (time(13, 0) <= start < time(13, 45))
    # Inter-class break (5 minutes)
    # This is handled in the scheduling logic, not here
    return morning_break or lunch_break

def check_professor_availability(professor_schedule, faculty, day, start_slot, duration, activity_type):
    """Check if a professor can be scheduled for a new class considering REQ-10"""
    # If this is a lab scheduled after a lecture/tutorial, we can allow it without time gap
    # We'd need to know the previous activity type, which isn't implemented yet
    
    # Check for any existing slots for this professor on this day
    existing_slots = sorted(list(professor_schedule[faculty][day]))
    if not existing_slots:
        return True  # No other classes on this day
        
    # Calculate end slot of proposed new class
    end_slot = start_slot + duration - 1
    
    # Check the minimum time difference between existing and new slots
    MIN_GAP_SLOTS = 6  # 3 hours = 6 half-hour slots
    
    for slot in existing_slots:
        # If existing slot is within or adjacent to requested slot range
        if start_slot <= slot <= end_slot:
            return False  # Direct conflict
            
        # Check if gap between classes is sufficient
        if slot < start_slot and start_slot - slot < MIN_GAP_SLOTS:
            # Existing class ends too close to new class
            return False
            
        if slot > end_slot and slot - end_slot < MIN_GAP_SLOTS:
            # New class ends too close to existing class
            return False
    
    return True

def check_professor_constraint(professor_schedule, faculty, day, start_slot, duration, timetable, time_slots):
    """Check if a professor can be scheduled for a new class considering REQ-10"""
    # If professor has no classes that day, constraint is satisfied
    if not professor_schedule[faculty][day]:
        return True
    
    # Get the new class's time range
    new_class_start_time = time_slots[start_slot][0]
    new_class_end_time = time_slots[start_slot + duration - 1][1]
    
    # Create datetime objects for today to allow time subtraction
    today = datetime.today().date()
    new_start_datetime = datetime.combine(today, new_class_start_time)
    new_end_datetime = datetime.combine(today, new_class_end_time)
    
    # Check each existing class for this professor on this day
    for existing_slot in professor_schedule[faculty][day]:
        # Get information about the existing class
        existing_class_type = timetable[day][existing_slot]['type']
        
        # Skip slots that don't have class type (might be continuation slots)
        if existing_class_type is None:
            continue
        
        # Get existing class time
        existing_class_start_time = time_slots[existing_slot][0]
        existing_class_end_time = None
        
        # Find the end time by looking for the last slot of this class
        for i in range(existing_slot, len(time_slots)):
            if i in professor_schedule[faculty][day] and timetable[day][i]['type'] is not None:
                existing_class_end_time = time_slots[i][1]
            else:
                break
        
        # Skip if we couldn't determine the end time
        if existing_class_end_time is None:
            continue
        
        # Convert to datetime objects for comparisons
        existing_start_datetime = datetime.combine(today, existing_class_start_time)
        existing_end_datetime = datetime.combine(today, existing_class_end_time)
        
        # Special case: Lab can be scheduled after lecture/tutorial
        if (new_start_datetime == existing_end_datetime and 
            existing_class_type in ['LEC', 'TUT'] and duration == LAB_DURATION):
            continue
        
        if (existing_start_datetime == new_end_datetime and 
            timetable[day][start_slot]['type'] == 'LAB' and existing_class_type in ['LEC', 'TUT']):
            continue
        
        # Calculate time difference in hours
        time_diff_hours = abs((new_start_datetime - existing_start_datetime).total_seconds() / 3600)
        
        # Check if classes are consecutive (not allowed unless it's the special case)
        if new_start_datetime == existing_end_datetime or existing_start_datetime == new_end_datetime:
            return False
        
        # Check if the time difference is less than 3 hours
        if time_diff_hours < 3:
            return False
    
    return True

def generate_all_timetables():
    TIME_SLOTS = generate_time_slots()
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    professor_schedule = {}   # Track professor assignments
    classroom_schedule = {}   # Track classroom assignments
    
    # Initialize classroom schedules for all rooms
    for room in lecture_rooms + computer_lab_rooms + large_rooms:
        classroom_schedule[room] = {day: set() for day in range(len(DAYS))}
    
    # Pre-assign a fixed lecture room for each department/semester combination
    section_lecture_rooms = {}
    section_lab_rooms = {}
    available_lecture_rooms = lecture_rooms.copy()
    available_lab_rooms = computer_lab_rooms.copy()
    
    # First identify all department/semester combinations
    dept_sem_combinations = []
    for department in df['Department'].unique():
        for semester in df[df['Department'] == department]['Semester'].unique():
            dept_sem_combinations.append((department, semester))
    
    # Randomly assign one lecture room to each department/semester
    for dept, sem in dept_sem_combinations:
        section_key = (dept, sem)
        if available_lecture_rooms:
            # Pop a room from available rooms to ensure no duplicates
            assigned_room = available_lecture_rooms.pop(0)
            section_lecture_rooms[section_key] = assigned_room
            # Re-add at the end to ensure rooms are reused only after all others are used
            available_lecture_rooms.append(assigned_room)
        else:
            # If we run out of lecture rooms, start reusing from the beginning
            section_lecture_rooms[section_key] = lecture_rooms[0] if lecture_rooms else "No Lecture Room"
            
        # Assign two lab rooms for lab sessions to accommodate more students
        if len(computer_lab_rooms) >= 2:
            # Take two lab rooms if available
            assigned_lab1 = computer_lab_rooms[0]
            assigned_lab2 = computer_lab_rooms[1]
            # Store as a list of two labs
            section_lab_rooms[section_key] = [assigned_lab1, assigned_lab2]
        elif computer_lab_rooms:
            # If only one lab room is available, use it twice (though this is not ideal)
            assigned_lab = computer_lab_rooms[0]
            section_lab_rooms[section_key] = [assigned_lab, assigned_lab]
        else:
            # If no lab rooms available
            section_lab_rooms[section_key] = ["No Lab Room", "No Lab Room"]
    
    # Create a dictionary to store assigned classrooms for each course
    course_classrooms = {}
    
    # Identify courses with multiple sections in the same semester
    section_counts = df.groupby(['Department', 'Semester', 'Course Code']).size().to_dict()
    multi_section_courses = {k: v for k, v in section_counts.items() if v > 1}
    
    for department in df['Department'].unique():
        for semester in df[df['Department'] == department]['Semester'].unique():
            courses = df[(df['Department'] == department) & (df['Semester'] == semester)].copy()
            
            if courses.empty:
                continue
            
            # Create worksheet for this department-semester
            ws = wb.create_sheet(title=f"{department}_{semester}")
            
            # Initialize timetable structure
            timetable = {day: {slot: {'type': None, 'code': '', 'name': '', 'faculty': '', 'classroom': ''} 
                         for slot in range(len(TIME_SLOTS))} for day in range(len(DAYS))}
            
            # Dictionary to store course colors and generate course color generator
            course_colors = {}
            color_gen = generate_course_color()
            
            # Get the pre-assigned lecture room for this section
            section_key = (department, semester)
            assigned_lecture_room = section_lecture_rooms.get(section_key, "No Lecture Room")
            assigned_lab_rooms = section_lab_rooms.get(section_key, ["No Lab Room", "No Lab Room"]) 
            
            # Process all courses to ensure LTPS requirements are met
            for _, course in courses.iterrows():
                code = str(course['Course Code'])
                name = str(course['Course Name'])
                faculty = str(course['Faculty'])
                
                # Extract LTPS values, ensuring they're integers or 0 if NaN
                l_hours = int(course['L']) if pd.notna(course['L']) else 0
                t_hours = int(course['T']) if pd.notna(course['T']) else 0
                p_hours = int(course['P']) if pd.notna(course['P']) else 0
                s_hours = int(course['S']) if pd.notna(course['S']) and 'S' in course else 0
                
                # Skip if no hours are scheduled for this course
                if l_hours == 0 and t_hours == 0 and p_hours == 0:
                    continue
                
                # Assign a color to this course if not already assigned
                if code not in course_colors:
                    # Store LTPS values with the course info
                    course_colors[code] = {
                        "color": next(color_gen), 
                        "name": name, 
                        "faculty": faculty,
                        "L": l_hours,
                        "T": t_hours,
                        "P": p_hours,
                        "S": s_hours
                    }
                
                if faculty not in professor_schedule:
                    professor_schedule[faculty] = {day: set() for day in range(len(DAYS))}
                
                # Calculate number of sessions needed
                # For lectures: 1.5-hour sessions
                num_lecture_sessions = l_hours // 1.5
                if l_hours % 1.5 > 0:
                    num_lecture_sessions += 1
                
                # For tutorials: 1-hour sessions 
                num_tutorial_sessions = t_hours
                
                # For practicals: 2-hour sessions
                num_lab_sessions = p_hours // 2
                if p_hours % 2 > 0:
                    num_lab_sessions += 1
                
                # Store which days course components have been scheduled on
                scheduled_days = set()
                
                # Schedule practicals first (if any)
                if p_hours > 0:
                    for _ in range(int(num_lab_sessions)):
                        # Try to schedule on days where no other component of this course exists
                        scheduled = False
                        attempts = 0
                        preferred_days = [d for d in range(len(DAYS)) if d not in scheduled_days]
                        
                        # If all days already have components, allow scheduling on any day
                        if not preferred_days:
                            preferred_days = list(range(len(DAYS)))
                        
                        while not scheduled and attempts < 1000:
                            # Prioritize preferred days but fallback to any day if needed
                            if preferred_days and attempts < 500:
                                day = random.choice(preferred_days)
                            else:
                                day = random.randint(0, len(DAYS)-1)
                            
                            if len(TIME_SLOTS) >= LAB_DURATION:
                                start_slot = random.randint(0, len(TIME_SLOTS)-LAB_DURATION)
                                
                                # Check if all required slots are free and not in break time
                                slots_free = True
                                for i in range(LAB_DURATION):
                                    if start_slot+i in professor_schedule[faculty][day] or is_break_time(TIME_SLOTS[start_slot+i]):
                                        slots_free = False
                                        break
                                    
                                    # Check classroom conflicts - handle both single and dual lab cases
                                    if isinstance(assigned_lab_rooms, list):
                                        lab1, lab2 = assigned_lab_rooms
                                        if (start_slot+i in classroom_schedule[lab1][day] or 
                                            start_slot+i in classroom_schedule[lab2][day] or
                                            timetable[day][start_slot+i]['type'] is not None):
                                            slots_free = False
                                            break
                                    else:
                                        if (start_slot+i in classroom_schedule[assigned_lab_rooms][day] or
                                            timetable[day][start_slot+i]['type'] is not None):
                                            slots_free = False
                                            break
                                
                                # Add check for professor constraint (REQ-10)
                                if slots_free:
                                    slots_free = check_professor_constraint(professor_schedule, faculty, day, start_slot, LAB_DURATION, timetable, TIME_SLOTS)
                                
                                if slots_free:
                                    # Mark professor and classroom(s) as busy
                                    for i in range(LAB_DURATION):
                                        professor_schedule[faculty][day].add(start_slot+i)
                                        
                                        # Update classroom schedules based on type
                                        if isinstance(assigned_lab_rooms, list):
                                            lab1, lab2 = assigned_lab_rooms
                                            classroom_schedule[lab1][day].add(start_slot+i)
                                            classroom_schedule[lab2][day].add(start_slot+i)
                                        else:
                                            classroom_schedule[assigned_lab_rooms][day].add(start_slot+i)
                                        
                                        timetable[day][start_slot+i]['type'] = 'LAB'
                                        timetable[day][start_slot+i]['code'] = code if i == 0 else ''
                                        timetable[day][start_slot+i]['name'] = name if i == 0 else ''
                                        timetable[day][start_slot+i]['faculty'] = faculty if i == 0 else ''
                                        timetable[day][start_slot+i]['classroom'] = f"{assigned_lab_rooms[0]} & {assigned_lab_rooms[1]}" if i == 0 else ''
                                    scheduled = True
                                    scheduled_days.add(day)
                            attempts += 1
                
                # Schedule lectures
                if l_hours > 0:
                    for _ in range(int(num_lecture_sessions)):
                        # Try to schedule on days where no other component of this course exists
                        scheduled = False
                        attempts = 0
                        preferred_days = [d for d in range(len(DAYS)) if d not in scheduled_days]
                        
                        # If all days already have components, allow scheduling on any day
                        if not preferred_days:
                            preferred_days = list(range(len(DAYS)))
                        
                        while not scheduled and attempts < 1000:
                            # Prioritize preferred days but fallback to any day if needed
                            if preferred_days and attempts < 500:
                                day = random.choice(preferred_days)
                            else:
                                day = random.randint(0, len(DAYS)-1)
                                
                            if len(TIME_SLOTS) >= LECTURE_DURATION:
                                start_slot = random.randint(0, len(TIME_SLOTS)-LECTURE_DURATION)
                                
                                # Check if all required slots are free and not in break time
                                slots_free = True
                                for i in range(LECTURE_DURATION):
                                    if (start_slot+i in professor_schedule[faculty][day] or 
                                        start_slot+i in classroom_schedule[assigned_lecture_room][day] or
                                        timetable[day][start_slot+i]['type'] is not None or
                                        is_break_time(TIME_SLOTS[start_slot+i])):
                                        slots_free = False
                                        break
                                
                                # Add check for professor constraint (REQ-10)
                                if slots_free:
                                    slots_free = check_professor_constraint(professor_schedule, faculty, day, start_slot, LECTURE_DURATION, timetable, TIME_SLOTS)
                                
                                if slots_free:
                                    # Mark professor and classroom as busy
                                    for i in range(LECTURE_DURATION):
                                        professor_schedule[faculty][day].add(start_slot+i)
                                        classroom_schedule[assigned_lecture_room][day].add(start_slot+i)
                                        timetable[day][start_slot+i]['type'] = 'LEC'
                                        timetable[day][start_slot+i]['code'] = code if i == 0 else ''
                                        timetable[day][start_slot+i]['name'] = name if i == 0 else ''
                                        timetable[day][start_slot+i]['faculty'] = faculty if i == 0 else ''
                                        timetable[day][start_slot+i]['classroom'] = assigned_lecture_room if i == 0 else ''
                                    scheduled = True
                                    scheduled_days.add(day)
                            attempts += 1
                
                # Schedule tutorials
                if t_hours > 0:
                    for _ in range(int(num_tutorial_sessions)):
                        scheduled = False
                        attempts = 0
                        while not scheduled and attempts < 1000:
                            day = random.randint(0, len(DAYS)-1)
                            if len(TIME_SLOTS) >= TUTORIAL_DURATION:
                                start_slot = random.randint(0, len(TIME_SLOTS)-TUTORIAL_DURATION)
                                
                                # Skip if it's break time
                                if is_break_time(TIME_SLOTS[start_slot]):
                                    attempts += 1
                                    continue
                                    
                                # Check if all required slots are free
                                slots_free = True
                                for i in range(TUTORIAL_DURATION):
                                    if (start_slot+i in professor_schedule[faculty][day] or 
                                        start_slot+i in classroom_schedule[assigned_lecture_room][day] or
                                        timetable[day][start_slot+i]['type'] is not None):
                                        slots_free = False
                                        break
                                
                                # Add check for professor constraint (REQ-10)
                                if slots_free:
                                    slots_free = check_professor_constraint(professor_schedule, faculty, day, start_slot, TUTORIAL_DURATION, timetable, TIME_SLOTS)
                                        
                                if slots_free:
                                    # Mark professor and classroom as busy
                                    for i in range(TUTORIAL_DURATION):
                                        professor_schedule[faculty][day].add(start_slot+i)
                                        classroom_schedule[assigned_lecture_room][day].add(start_slot+i)
                                        timetable[day][start_slot+i]['type'] = 'TUT'
                                        timetable[day][start_slot+i]['code'] = code if i == 0 else ''
                                        timetable[day][start_slot+i]['name'] = name if i == 0 else ''
                                        timetable[day][start_slot+i]['faculty'] = faculty if i == 0 else ''
                                        timetable[day][start_slot+i]['classroom'] = assigned_lecture_room if i == 0 else ''
                                    scheduled = True
                                    scheduled_days.add(day)
                            attempts += 1
            
            # Write timetable to worksheet with merged cells and breaks
            # Create header
            header = ['Day'] + [f"{slot[0].strftime('%H:%M')}-{slot[1].strftime('%H:%M')}" for slot in TIME_SLOTS]
            ws.append(header)
            
            # Apply header formatting
            header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            header_font = Font(bold=True)
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Fill data and merge cells
            break_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            border = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))
            
            for day_idx, day in enumerate(DAYS):
                row_num = day_idx + 2  # +1 for header, +1 because rows start at 1
                ws.append([day])
                
                # Track merged regions
                merge_ranges = []
                current_merge = None
                
                for slot_idx in range(len(TIME_SLOTS)):
                    cell_value = ''
                    cell_fill = None
                    
                    # Check if this is break time
                    if is_break_time(TIME_SLOTS[slot_idx]):
                        cell_value = "BREAK"
                        cell_fill = break_fill
                    elif timetable[day_idx][slot_idx]['type']:
                        if timetable[day_idx][slot_idx]['code']:  # First slot of activity
                            activity_type = timetable[day_idx][slot_idx]['type']
                            code = timetable[day_idx][slot_idx]['code']
                            classroom = timetable[day_idx][slot_idx]['classroom']
                            
                            # Use course-specific color
                            if code in course_colors:
                                cell_fill = PatternFill(start_color=course_colors[code]["color"], 
                                                        end_color=course_colors[code]["color"], 
                                                        fill_type="solid")
                            
                            if activity_type == 'LEC':
                                duration = LECTURE_DURATION
                            elif activity_type == 'LAB':
                                duration = LAB_DURATION
                            else:  # TUT
                                duration = TUTORIAL_DURATION
                            
                            # Create merged range
                            start_col = get_column_letter(slot_idx + 2)  # +1 for day column
                            end_col = get_column_letter(slot_idx + duration + 1)
                            merge_range = f"{start_col}{row_num}:{end_col}{row_num}"
                            merge_ranges.append(merge_range)
                            
                            # Display course code, activity type, and classroom in the timetable
                            cell_value = f"{code} {activity_type}\n room no. :{classroom}"
                
                    # Write to cell
                    cell = ws.cell(row=row_num, column=slot_idx+2, value=cell_value)
                    if cell_fill:
                        cell.fill = cell_fill
                    cell.border = border
                    cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                
                # Apply all merges for this row
                for merge_range in merge_ranges:
                    ws.merge_cells(merge_range)
            
            # Add a cell with information about assigned rooms for this section
            info_row = len(DAYS) + 3
            info_cell = ws.cell(row=info_row, column=1, value=f"Default Lecture Room: {assigned_lecture_room}")
            info_cell.font = Font(bold=True)
            
            # Display both assigned lab rooms
            lab_info = f"Default Lab Rooms: {assigned_lab_rooms[0]} & {assigned_lab_rooms[1]}"
            lab_info_cell = ws.cell(row=info_row+1, column=1, value=lab_info)
            lab_info_cell.font = Font(bold=True)
            
            # Adjust column widths and row heights for better visibility
            for col_idx in range(1, len(TIME_SLOTS)+2):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = 20  # Increase width from 15 to 20
            
            # Increase day column width for better visibility
            ws.column_dimensions['A'].width = 12
            
            # Increase row heights for proper text display
            for row in ws.iter_rows(min_row=2, max_row=len(DAYS)+1):
                ws.row_dimensions[row[0].row].height = 60  # Increase height from 40 to 60
            
            # Add extra padding between timetable and legend
            legend_start_row = len(DAYS) + 8  # Increased spacing before legend
            
            # Create a section header for the legend with better visibility
            legend_header = ws.cell(row=legend_start_row, column=1, value="LEGEND")
            legend_header.font = Font(bold=True, size=16)  # Increased font size
            legend_header.alignment = Alignment(horizontal='left', vertical='center')
            ws.row_dimensions[legend_start_row].height = 35  # Taller row for header
            
            # Add background shading for the legend header row with wider span
            for col_idx in range(1, 6):  # Extended to 5 columns
                cell = ws.cell(row=legend_start_row, column=col_idx)
                cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # Lighter blue
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                   top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Format break section with better spacing
            break_row = legend_start_row + 3  # Added more spacing after header
            ws.cell(row=break_row, column=1, value="BREAK TIME:").font = Font(bold=True, size=12)
            
            break_cell = ws.cell(row=break_row, column=2, value="")
            break_cell.fill = break_fill
            break_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                     top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Add break time descriptions with updated times
            break_info = ws.cell(row=break_row, column=3, value="Morning: 10:30-10:45 (15 min), Lunch: 13:00-13:45 (45 min), Inter-class: 5 min")
            break_info.alignment = Alignment(horizontal='left', vertical='center')
            break_info.font = Font(size=11)
            ws.row_dimensions[break_row].height = 28
            
            # Add course legend with colors and better formatting
            legend_row = break_row + 4  # Extra spacing after break section
            
            # Add a section header for courses
            course_header = ws.cell(row=legend_row-1, column=1, value="COURSES:")
            course_header.font = Font(bold=True, size=12)
            ws.row_dimensions[legend_row-1].height = 28
            
            # Format header row of legend with better visibility
            headers = ["Course Code", "Color", "Course Name", "Faculty", "L-T-P-S"]
            for idx, header in enumerate(headers):
                header_cell = ws.cell(row=legend_row, column=idx+1, value=header)
                header_cell.font = Font(bold=True)
                header_cell.fill = PatternFill(start_color="D8E4BC", end_color="D8E4BC", fill_type="solid")  # Softer green
                header_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                         top=Side(style='thin'), bottom=Side(style='thin'))
                header_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Set appropriate widths for all legend columns
            ws.column_dimensions[get_column_letter(1)].width = 18  # Course code column - increased width
            ws.column_dimensions[get_column_letter(2)].width = 15  # Color column - increased width
            ws.column_dimensions[get_column_letter(3)].width = 45  # Name column - wider for long course names
            ws.column_dimensions[get_column_letter(4)].width = 35  # Faculty column - wider for longer names
            ws.column_dimensions[get_column_letter(5)].width = 15  # LTPS column
            
            # Add course details with proper formatting and padding
            for i, (code, details) in enumerate(course_colors.items()):
                row = legend_row + i + 1
                ws.row_dimensions[row].height = 60  # Increased height for better readability
                
                # Course code cell
                code_cell = ws.cell(row=row, column=1, value=code)
                code_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                       top=Side(style='thin'), bottom=Side(style='thin'))
                code_cell.alignment = Alignment(horizontal='center', vertical='center')
                code_cell.font = Font(size=11)  # Added standard font size
                
                # Color cell
                color_cell = ws.cell(row=row, column=2, value="")
                color_cell.fill = PatternFill(start_color=details["color"], 
                                            end_color=details["color"], 
                                            fill_type="solid")
                color_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                        top=Side(style='thin'), bottom=Side(style='thin'))
                
                # Course name cell with extra padding
                name_cell = ws.cell(row=row, column=3, value=details["name"])
                name_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                       top=Side(style='thin'), bottom=Side(style='thin'))
                name_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=2)  # Increased indent
                name_cell.font = Font(size=11)  # Added standard font size
                
                # Faculty cell with extra padding
                faculty_cell = ws.cell(row=row, column=4, value=details["faculty"])
                faculty_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                          top=Side(style='thin'), bottom=Side(style='thin'))
                faculty_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=2)  # Increased indent
                faculty_cell.font = Font(size=11)  # Added standard font size
                
                # Add LTPS cell
                ltps_text = f"{details.get('L', 0)}-{details.get('T', 0)}-{details.get('P', 0)}-{details.get('S', 0)}"
                ltps_cell = ws.cell(row=row, column=5, value=ltps_text)
                ltps_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                       top=Side(style='thin'), bottom=Side(style='thin'))
                ltps_cell.alignment = Alignment(horizontal='center', vertical='center')
                ltps_cell.font = Font(size=11)
                
                # Add alternating row colors for better readability
                if i % 2 == 1:
                    for col_idx in range(1, 6):  # Updated to include LTPS column
                        if col_idx != 2:  # Skip the color cell
                            cell = ws.cell(row=row, column=col_idx)
                            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Lighter gray
            
            # Add some padding at the bottom of the legend
            ws.row_dimensions[legend_row + len(course_colors) + 1].height = 20
    
    # Save the workbook
    wb.save("timetables.xlsx")
    print("Final timetables with breaks and course legend saved to timetables.xlsx")

def check_unscheduled_courses():
    """Check and print courses that are not scheduled according to their L-T-P-S requirements"""
    try:
        # Load the original course requirements
        df = pd.read_excel('combined.xlsx', sheet_name='Sheet1')
        
        # Load the generated timetable
        wb = pd.ExcelFile('timetables.xlsx')
        
        # Dictionary to track scheduled hours for each course
        scheduled_hours = defaultdict(lambda: {'L': 0, 'T': 0, 'P': 0, 'S': 0})
        
        # Process each sheet in the timetable workbook to extract all scheduled courses
        found_courses = []
        
        print("\nExamining timetable for scheduled courses...")
        
        for sheet_name in wb.sheet_names:
            timetable_df = pd.read_excel(wb, sheet_name=sheet_name)
            
            # Skip if this is not a valid timetable sheet (should have at least 'Day' column)
            if 'Day' not in timetable_df.columns:
                continue
            
            print(f"Processing sheet: {sheet_name}")
            
            # Process each row (day) in the timetable
            for _, row in timetable_df.iterrows():
                day = row['Day']
                if day not in DAYS:  # Skip header or legend rows
                    continue
                
                # Process each time slot in the day
                for col in timetable_df.columns[1:]:  # Skip the 'Day' column
                    cell_value = row[col]
                    
                    # Skip empty cells or break cells
                    if pd.isna(cell_value) or cell_value == '' or cell_value == 'BREAK':
                        continue
                    
                    # Process cell content to extract course information
                    # Format example: "CS101 LEC room no. :A101"
                    if isinstance(cell_value, str) and 'room no.' in cell_value:
                        # Split the cell value to extract components
                        parts = cell_value.split('room no.')
                        course_info = parts[0].strip()
                        
                        # Extract course code and class type
                        if ' ' in course_info:
                            course_parts = course_info.split()
                            if len(course_parts) >= 2:
                                course_code = course_parts[0].strip()
                                class_type = course_parts[1].strip()
                                
                                # Debug output for problematic courses
                                if 'HS204' in course_code or 'HS153' in course_code:
                                    print(f"Found in timetable: {course_code}, Type: {class_type}, Day: {day}, Slot: {col}")
                                
                                found_courses.append(course_code)  # Track all courses found in timetable
                                
                                # Update scheduled hours based on class type
                                if class_type == 'LEC':
                                    scheduled_hours[course_code]['L'] += 1.5  # Lecture is 1.5 hours
                                    print(f"Added 1.5 lecture hours for {course_code}")
                                elif class_type == 'TUT':
                                    scheduled_hours[course_code]['T'] += 1    # Tutorial is 1 hour
                                    print(f"Added 1 tutorial hour for {course_code}")
                                elif class_type == 'LAB':
                                    scheduled_hours[course_code]['P'] += 2    # Lab is 2 hours
                                    print(f"Added 2 practical hours for {course_code}")
        
        # Create mappings for course codes with aliases or variations
        all_found_courses = set(found_courses)
        course_primary_codes = {}  # Maps each course code variant to its primary code
        
        # Build a mapping of all course code variations
        for _, course in df.iterrows():
            original_code = str(course['Course Code']).strip()
            
            # For courses with slashes (e.g., "HS204 / HS153")
            if '/' in original_code:
                variants = [c.strip() for c in original_code.split('/')]
                primary = variants[0]
                
                # Map all variants to the primary code
                for variant in variants:
                    course_primary_codes[variant] = primary
                
                # Also map the complete original code
                course_primary_codes[original_code] = primary
                
                # Debug output for specific courses
                if 'HS204' in original_code or 'HS153' in original_code:
                    print(f"Mapping variants for {original_code}: {variants}")
            
            # For courses with parentheses (e.g., "B1(ASD151/HS151/New/New/New/New)")
            elif '(' in original_code and ')' in original_code:
                base = original_code.split('(')[0].strip()
                inner_part = original_code.split('(')[1].split(')')[0]
                
                if '/' in inner_part:
                    inner_variants = [c.strip() for c in inner_part.split('/')]
                    for variant in inner_variants:
                        if variant.lower() != 'new':
                            # Create combined codes like "B1_ASD151"
                            combined = f"{base}_{variant}"
                            course_primary_codes[combined] = original_code
                            course_primary_codes[variant] = original_code
                
                # Map the original code to itself
                course_primary_codes[original_code] = original_code
            
            # Regular codes just map to themselves
            else:
                course_primary_codes[original_code] = original_code
        
        # Print courses found in timetable for debugging
        print("\nUnique courses found in timetable:", len(all_found_courses))
        print("First 10 courses found:", list(all_found_courses)[:10])
        
        # Merge scheduled hours for course variants
        merged_hours = defaultdict(lambda: {'L': 0, 'T': 0, 'P': 0, 'S': 0})
        
        for code, hours in scheduled_hours.items():
            # Find the primary code (or use the code itself if not a variant)
            primary_code = course_primary_codes.get(code, code)
            
            # Debug output for problematic courses
            if 'HS204' in code or 'HS153' in code or 'HS204' in primary_code or 'HS153' in primary_code:
                print(f"Merging hours for {code} -> {primary_code}: {hours}")
            
            # Accumulate hours by primary code
            merged_hours[primary_code]['L'] += hours['L']
            merged_hours[primary_code]['T'] += hours['T']
            merged_hours[primary_code]['P'] += hours['P']
            merged_hours[primary_code]['S'] += hours['S']
        
        # Compare with required hours to find unscheduled courses
        unscheduled_courses = []
        
        for _, course in df.iterrows():
            original_code = str(course['Course Code']).strip()
            name = str(course['Course Name'])
            faculty = str(course['Faculty'])
            department = str(course['Department'])
            semester = str(course['Semester'])
            
            # Get the primary code for looking up scheduled hours
            primary_code = course_primary_codes.get(original_code, original_code)
            
            # Extract LTPS requirements
            required_l = int(course['L']) if pd.notna(course['L']) else 0
            required_t = int(course['T']) if pd.notna(course['T']) else 0
            required_p = int(course['P']) if pd.notna(course['P']) else 0
            required_s = int(course['S']) if pd.notna(course['S']) and 'S' in course else 0
            
            # Get scheduled hours for this course using the primary code
            scheduled_l = merged_hours[primary_code]['L']
            scheduled_t = merged_hours[primary_code]['T']
            scheduled_p = merged_hours[primary_code]['P']
            scheduled_s = merged_hours[primary_code]['S']
            
            # Debug output for specific courses
            if 'HS204' in original_code or 'HS153' in original_code:
                print(f"\nCourse: {original_code} (Primary: {primary_code})")
                print(f"  Required L-T-P-S: {required_l}-{required_t}-{required_p}-{required_s}")
                print(f"  Merged Scheduled L-T-P-S: {scheduled_l}-{scheduled_t}-{scheduled_p}-{scheduled_s}")
            
            # Use a tolerance for floating-point comparisons
            tolerance = 0.01
            
            # Check if any component is under-scheduled
            missing_l = max(0, required_l - scheduled_l)
            missing_t = max(0, required_t - scheduled_t)
            missing_p = max(0, required_p - scheduled_p)
            missing_s = max(0, required_s - scheduled_s)
            
            if (missing_l > tolerance or missing_t > tolerance or 
                missing_p > tolerance or missing_s > tolerance):
                
                # Determine possible reasons for unscheduled components
                reasons = []
                
                # Check if any variant of this course was found in the timetable
                variants_found = False
                found_variants = []
                
                # Check all variants
                if primary_code in course_primary_codes.values():
                    for code, primary in course_primary_codes.items():
                        if primary == primary_code and code in all_found_courses:
                            variants_found = True
                            found_variants.append(code)
                else:
                    # Check the primary code itself
                    if primary_code in all_found_courses:
                        variants_found = True
                        found_variants.append(primary_code)
                
                if variants_found:
                    reasons.append(f"Course found in timetable as {', '.join(found_variants)} but not all required hours are scheduled")
                else:
                    reasons.append("Course not found in any timetable")
                
                # Check for faculty conflicts
                faculty_courses = df[df['Faculty'] == faculty]['Course Code'].tolist()
                if len(faculty_courses) > 1:
                    reasons.append("Faculty teaching multiple courses may have scheduling constraints")
                
                # Check for room availability
                section_key = (department, semester)
                if len(lecture_rooms) < 1 and required_l > 0:
                    reasons.append("Insufficient lecture rooms available")
                if len(computer_lab_rooms) < 1 and required_p > 0:
                    reasons.append("Insufficient lab rooms available")
                
                # Check for semester course load
                semester_courses = df[(df['Department'] == department) & (df['Semester'] == semester)].shape[0]
                if semester_courses > 6:
                    reasons.append("High number of courses in same semester may cause conflicts")
                
                # Add course to unscheduled list
                unscheduled_courses.append({
                    'Code': original_code,
                    'Name': name,
                    'Faculty': faculty,
                    'Department': department,
                    'Semester': semester,
                    'Required L-T-P-S': f"{required_l}-{required_t}-{required_p}-{required_s}",
                    'Scheduled L-T-P-S': f"{scheduled_l}-{scheduled_t}-{scheduled_p}-{scheduled_s}",
                    'Missing L': round(missing_l, 2),
                    'Missing T': round(missing_t, 2),
                    'Missing P': round(missing_p, 2),
                    'Missing S': round(missing_s, 2),
                    'Variant Found': variants_found,
                    'Found As': ', '.join(found_variants) if found_variants else "Not found",
                    'Reasons': "; ".join(reasons)
                })
        
        # Print results
        if unscheduled_courses:
            print("\n=== COURSES WITH UNSCHEDULED HOURS ===")
            print(f"Found {len(unscheduled_courses)} courses with scheduling issues:\n")
            
            for course in unscheduled_courses:
                print(f"Course: {course['Code']} - {course['Name']}")
                print(f"  Department: {course['Department']}, Semester: {course['Semester']}")
                print(f"  Faculty: {course['Faculty']}")
                print(f"  Required L-T-P-S: {course['Required L-T-P-S']}")
                print(f"  Scheduled L-T-P-S: {course['Scheduled L-T-P-S']}")
                
                missing = []
                if course['Missing L'] > 0:
                    missing.append(f"{course['Missing L']} lecture hours")
                if course['Missing T'] > 0:
                    missing.append(f"{course['Missing T']} tutorial hours")
                if course['Missing P'] > 0:
                    missing.append(f"{course['Missing P']} practical hours")
                if course['Missing S'] > 0:
                    missing.append(f"{course['Missing S']} self-study hours")
                    
                print(f"  Missing: {', '.join(missing)}")
                print(f"  Found in Timetable: {'Yes' if course['Variant Found'] else 'No'}")
                if course['Variant Found']:
                    print(f"  Found as: {course['Found As']}")
                print(f"  Possible Reasons: {course['Reasons']}\n")
            
            # Create Excel file with unscheduled courses
            unscheduled_df = pd.DataFrame(unscheduled_courses)
            unscheduled_df.to_excel('unscheduled_courses.xlsx', index=False)
            print("Details saved to 'unscheduled_courses.xlsx'")
        else:
            print("\n=== ALL COURSES FULLY SCHEDULED ===")
            print("All courses have been scheduled according to their L-T-P-S requirements.")
            
    except Exception as e:
        print(f"Error checking unscheduled courses: {e}")
        import traceback
        traceback.print_exc()

def generate_faculty_timetables():
    """Generate individual timetables for each faculty member"""
    try:
        # Load the generated timetable
        wb = pd.ExcelFile('timetables.xlsx')
        
        # Dictionary to track faculty schedules
        faculty_schedules = {}
        
        # Process each sheet in the timetable workbook
        for sheet_name in wb.sheet_names:
            timetable_df = pd.read_excel(wb, sheet_name=sheet_name)
            
            # Skip if this is not a valid timetable sheet (should have at least 'Day' column)
            if 'Day' not in timetable_df.columns:
                continue
                
            # Extract department and semester from sheet name
            dept_sem = sheet_name
                
            # Process each row (day) in the timetable
            for _, row in timetable_df.iterrows():
                day = row['Day']
                if day not in DAYS:  # Skip header or legend rows
                    continue
                    
                # Process each time slot in the day
                for col in timetable_df.columns[1:]:  # Skip the 'Day' column
                    cell_value = row[col]
                    
                    # Skip empty cells or break cells
                    if pd.isna(cell_value) or cell_value == '' or cell_value == 'BREAK':
                        continue
                        
                    # Extract course code, class type, and faculty
                    # Format example: "CS101 LEC room no. :A101"
                    if isinstance(cell_value, str) and 'room no.' in cell_value:
                        parts = cell_value.split()
                        if len(parts) >= 2:
                            # Find faculty information from combined.xlsx
                            course_code = parts[0]
                            class_type = parts[1]
                            
                            try:
                                # Get the original course data to extract faculty info
                                course_data = pd.read_excel('combined.xlsx', sheet_name='Sheet1')
                                course_row = course_data[course_data['Course Code'] == course_code]
                                
                                if not course_row.empty:
                                    faculty = str(course_row['Faculty'].iloc[0])
                                    room_info = cell_value.split('room no. :')[1].strip() if 'room no. :' in cell_value else "Unknown"
                                    course_name = str(course_row['Course Name'].iloc[0])
                                    
                                    # Initialize faculty entry if not already exists
                                    if faculty not in faculty_schedules:
                                        faculty_schedules[faculty] = {d: {} for d in DAYS}
                                    
                                    # Add this class to faculty schedule
                                    time_slot_str = col
                                    faculty_schedules[faculty][day][time_slot_str] = {
                                        'Course Code': course_code,
                                        'Course Name': course_name,
                                        'Class Type': class_type,
                                        'Room': room_info,
                                        'Department-Semester': dept_sem
                                    }
                            except Exception as e:
                                print(f"Error finding faculty for {course_code}: {e}")
        
        # Create a new workbook for faculty timetables
        faculty_wb = Workbook()
        faculty_wb.remove(faculty_wb.active)  # Remove default sheet
        
        # Sort faculty names for alphabetical order
        sorted_faculty_names = sorted(faculty_schedules.keys())
        
        # Define a function to sanitize sheet names (remove invalid Excel sheet name characters)
        def sanitize_sheet_name(name):
            # Replace characters that are not allowed in Excel sheet names
            invalid_chars = ['/', '\\', '?', '*', ':', '[', ']', "'"]
            sanitized_name = name
            for char in invalid_chars:
                sanitized_name = sanitized_name.replace(char, '_')
            
            # Excel sheet names are limited to 31 characters
            if len(sanitized_name) > 31:
                sanitized_name = sanitized_name[:31]
                
            return sanitized_name
        
        # Keep a mapping of sanitized names to original names to handle duplicates
        sanitized_to_original = {}
        
        for faculty in sorted_faculty_names:
            # Create a sanitized sheet name
            sanitized_faculty = sanitize_sheet_name(faculty)
            
            # Handle duplicate sanitized names by adding a number
            if sanitized_faculty in sanitized_to_original:
                base_name = sanitized_faculty
                counter = 1
                # Try adding numbers until we find a unique name
                while sanitized_faculty in sanitized_to_original:
                    sanitized_faculty = f"{base_name[:27]}_{counter}"  # Leave room for the counter
                    counter += 1
            
            # Store the mapping
            sanitized_to_original[sanitized_faculty] = faculty
            
            # Create a worksheet for this faculty using the sanitized name
            ws = faculty_wb.create_sheet(title=sanitized_faculty)
            
            # Add the original faculty name as the first row title
            ws.merge_cells('A1:G1')
            title_cell = ws['A1']
            title_cell.value = f"Schedule for: {faculty}"
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            title_cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            
            # Create header (starting from row 2 now)
            header = ['Day', 'Time Slot', 'Course Code', 'Course Name', 'Class Type', 'Room', 'Department-Semester']
            ws.append(header)
            
            # Apply header formatting
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            for cell in ws[2]:  # Now header is in row 2
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Add each scheduled class for this faculty
            row_idx = 3  # Start data from row 3
            for day in DAYS:
                # Sort time slots chronologically
                time_slots = sorted(faculty_schedules[faculty][day].keys())
                
                if not time_slots:  # No classes on this day
                    ws.append([day, "No classes scheduled", "", "", "", "", ""])
                    row_idx += 1
                    continue
                    
                for time_slot in time_slots:
                    class_info = faculty_schedules[faculty][day][time_slot]
                    
                    # Add this class to the worksheet
                    ws.append([
                        day,
                        time_slot,
                        class_info['Course Code'],
                        class_info['Course Name'],
                        class_info['Class Type'],
                        class_info['Room'],
                        class_info['Department-Semester']
                    ])
                    
                    # Apply formatting
                    for cell in ws[row_idx]:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                           top=Side(style='thin'), bottom=Side(style='thin'))
                    
                    # Highlight each class type differently
                    class_type = class_info['Class Type']
                    if class_type == 'LEC':
                        fill_color = "B8CCE4"  # Light blue
                    elif class_type == 'TUT':
                        fill_color = "E4B8CC"  # Pink
                    elif class_type == 'LAB':
                        fill_color = "CCE4B8"  # Light green
                    else:
                        fill_color = "F2F2F2"  # Light gray
                        
                    # Apply fill color
                    for cell in ws[row_idx]:
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    
                    row_idx += 1
            
            # Set predefined column widths instead of auto-adjusting to avoid MergedCell error
            column_widths = {
                'A': 15,  # Day
                'B': 20,  # Time Slot
                'C': 15,  # Course Code
                'D': 40,  # Course Name
                'E': 12,  # Class Type
                'F': 25,  # Room
                'G': 30,  # Department-Semester
            }
            
            # Apply the predefined widths
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width
        
        # Save the workbook
        faculty_wb.save("faculties_timetable.xlsx")
        print("Faculty timetables generated and saved to faculties_timetable.xlsx")
        
    except Exception as e:
        print(f"Error generating faculty timetables: {e}")
        # Print more detailed error information for debugging
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    generate_all_timetables()
    check_unscheduled_courses()
    generate_faculty_timetables()